from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
import os

app = Flask(__name__, static_folder='static')

EXCEL_PATH = os.path.join(os.getcwd(), "Jack's reservations.xlsx")

@app.route('/')
def index():
    # Assuming your HTML (with name jacks_dock_booking.html) lives in ./static
    return send_from_directory('static', 'jacks_dock_booking.html')

@app.route('/reserve', methods=['POST'])
def reserve():
    data = request.get_json()
    # expect keys: date, start, end, name, email, phone, cost
    required = ['date','start','end','name','email','phone','cost']
    if not all(k in data for k in required):
        return jsonify({ 'success': False, 'error': 'Missing field' }), 400

    # Load workbook and the first sheet
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Append header if sheet empty
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(['Date','Start','End','Name','Email','Phone','Cost'])

    # Append new reservation
    ws.append([
        data['date'],
        data['start'],
        data['end'],
        data['name'],
        data['email'],
        data['phone'],
        data['cost'],
    ])

    # Save back to disk
    wb.save(EXCEL_PATH)

    return jsonify({ 'success': True }), 200

if __name__ == '__main__':
    # run on localhost:5000
    app.run(debug=True)
